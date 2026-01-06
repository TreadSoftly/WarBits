#!/usr/bin/env python3
# pyright: reportUnknownArgumentType=false, reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnnecessaryCast=false
"""WarBits Excel ingestion (v1)

Reads the four War Thunder / WarBits Excel workbooks and writes normalized, domain-specific tables:

- vehicles.jsonl
- aircraft_performance_points_fc2.csv
- sea_level_speed_fc1.csv
- weapons.jsonl
- warheads.jsonl
- sensors_tgp_and_sights.jsonl
- custom_loadout_meta.jsonl
- custom_loadout_options.jsonl
- vehicle_alias_candidates.jsonl

This script is intentionally conservative:
- It preserves unknown/extra columns in `raw_fields`.
- It merges across sources without deleting partial records.

Run:
  python ingest_warbits_excels_v1.py \
    --factchecker1 WarThunder_FactChecker1.xlsx \
    --factchecker2 WarThunder_FactChecker2.xlsx \
    --factchecker3 WarThunder_FactChecker3.xlsx \
    --special specialialdata.xlsx \
    --out_dir warbits_normalized_v1
"""

from __future__ import annotations

import argparse
import datetime
import json
import math
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

import numpy as np
import openpyxl
import pandas as pd

FieldCoerce = type[str] | type[float]


def _is_na(value: object) -> bool:
    try:
        return bool(pd.isna(cast(Any, value)))
    except Exception:
        return False


def _is_not_na(value: object) -> bool:
    try:
        return bool(pd.notna(cast(Any, value)))
    except Exception:
        return False


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float, np.integer, np.floating)):
        num = float(value)
        return None if math.isnan(num) else num
    if isinstance(value, (bytes, bytearray, memoryview)):
        try:
            value = bytes(value).decode(errors="ignore")
        except Exception:
            return None
    s = str(value).strip()
    if s in {"", "-"}:
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def _to_int(value: object) -> Optional[int]:
    num = _to_float(value)
    if num is None:
        return None
    try:
        return int(num)
    except Exception:
        return None


def slugify(text: Any) -> str:
    s = str(text).strip().lower()
    s = s.replace('&', 'and')
    s = re.sub(r'[\(\)\[\]\{\}]', '', s)
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s


def parse_bool(val: Any) -> Optional[bool]:
    if val is None or _is_na(val):
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in {'yes', 'y', 'true', '1'}:
        return True
    if s in {'no', 'n', 'false', '0'}:
        return False
    return None


def parse_h_m_string(val: Any) -> Optional[float]:
    """Parse strings like '1h 31m' into seconds."""
    if val is None or _is_na(val):
        return None
    s = str(val).strip().lower()
    h = 0
    m = 0
    mh = re.search(r'(\d+)\s*h', s)
    if mh:
        h = int(mh.group(1))
    mm = re.search(r'(\d+)\s*m', s)
    if mm:
        m = int(mm.group(1))
    if h == 0 and m == 0:
        return None
    return h * 3600.0 + m * 60.0


def parse_mm_ss_cc(val: Any) -> Optional[float]:
    """Parse climb-time style values from the FactChecker2 sheets.

    In these sheets, Excel time values appear to be stored as:
      datetime.time(hour=<minutes>, minute=<seconds>, second=<centiseconds>)

    Example:
      datetime.time(1, 55) should be interpreted as 1m55s (115 seconds).
    """
    if val is None or _is_na(val):
        return None
    if isinstance(val, datetime.time):
        return val.hour * 60.0 + val.minute + val.second / 100.0
    if isinstance(val, datetime.timedelta):
        return val.total_seconds()
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in {'', '-'}:
        return None
    parts = [re.sub(r'[^0-9]', '', p) for p in s.split(':')]
    if any(p == '' for p in parts):
        return None
    nums = [int(p) for p in parts]
    if len(nums) == 2:
        m, sec = nums
        return m * 60.0 + sec
    if len(nums) == 3:
        m, sec, cc = nums
        return m * 60.0 + sec + cc / 100.0
    if len(nums) == 4:
        _, m, sec, cc = nums
        return m * 60.0 + sec + cc / 100.0
    return None


def make_json_safe(obj: Any) -> Any:
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        if isinstance(obj, float) and math.isnan(obj):
            return None
        return obj
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        v = float(obj)
        return None if math.isnan(v) else v
    if isinstance(obj, (datetime.date, datetime.datetime, datetime.time)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [make_json_safe(v) for v in obj]
    return str(obj)


def merge_dict(dst: Dict[str, Any], src: Dict[str, Any]) -> Dict[str, Any]:
    for k, v in src.items():
        if k not in dst or dst[k] is None:
            dst[k] = v
    return dst


def coerce_value(val: Any) -> Any:
    if val is None or _is_na(val):
        return None
    if isinstance(val, np.generic):
        val = val.item()
    if isinstance(val, str):
        s = val.strip()
        if s == '' or s == '-':
            return None
        b = parse_bool(s)
        if b is not None:
            return b
        s2 = s.replace(',', '')
        # keep ranges and zoom ratios as strings
        if re.match(r'^[+-]?\d+(\.\d+)?\s*/\s*[+-]?\d+(\.\d+)?$', s2):
            return s
        if 'x' in s.lower() or '/' in s:
            return s
        try:
            if '.' in s2 or 'e' in s2.lower():
                return float(s2)
            return int(s2)
        except Exception:
            return s
    return val


# -----------------------
# FactChecker2 (aircraft)
# -----------------------

def parse_aircraft_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    source_file: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    records: List[Dict[str, Any]] = []
    perf_points: List[Dict[str, Any]] = []

    name_rows = [
        i
        for i in range(2, len(df))
        if _is_not_na(df.iloc[i, 0]) and str(df.iloc[i, 0]).strip() not in {"", "nan"}
    ]
    name_rows_sorted = sorted(name_rows)

    def clean_num(v: Any) -> Optional[float]:
        return _to_float(v)

    for idx, start in enumerate(name_rows_sorted):
        end = name_rows_sorted[idx + 1] if idx + 1 < len(name_rows_sorted) else len(df)
        block = df.iloc[start:end].reset_index(drop=True)

        name = block.iloc[0, 0]
        if _is_na(name):
            continue
        name_str = str(name).strip()
        if name_str in {'Aircraft Details', 'Name'}:
            continue

        vehicle_id = slugify(name_str)
        rec: Dict[str, Any] = {
            "vehicle_id": vehicle_id,
            "name": name_str,
            "notes": None if _is_na(block.iloc[0, 1]) else str(block.iloc[0, 1]).strip(),
            "vehicle_type": "aircraft",
            "sources": [{
                "file": os.path.basename(source_file),
                "sheet": sheet_name,
                "row_start": int(start) + 1,
            }],
            "engine": {},
            "supercharger": {},
            "controls": {},
            "fuel": {},
            "performance": {
                "vy": {
                    "climb_speed_type": None,
                    "climb_speed_setting_1": None,
                    "climb_speed_wep": None,
                    "climb_time_curve": [],
                },
                "max_speed_wep_curve": [],
            },
        }

        # climb speed at Vy row
        climb_speed_type = block.iloc[0, 3]
        if _is_not_na(climb_speed_type):
            rec["performance"]["vy"]["climb_speed_type"] = str(climb_speed_type).strip()
        for key, col in [("climb_speed_setting_1", 4), ("climb_speed_wep", 5)]:
            v = block.iloc[0, col]
            if _is_not_na(v) and str(v).strip() not in {"-"}:
                val_num = _to_float(v)
                rec["performance"]["vy"][key] = val_num if val_num is not None else str(v).strip()

        # supercharger + hp
        rec["supercharger"]["stage2_switch_m"] = clean_num(block.iloc[0, 7])
        rec["supercharger"]["stage3_switch_m"] = clean_num(block.iloc[0, 8])
        rec["engine"]["stage2_hp"] = clean_num(block.iloc[0, 9])
        rec["engine"]["stage3_hp"] = clean_num(block.iloc[0, 10])

        # fuel
        fuel_val = block.iloc[0, 11] if block.shape[1] > 11 else None
        if _is_not_na(fuel_val):
            rec["fuel"]["full_fuel_duration_s"] = parse_h_m_string(fuel_val)
            rec["fuel"]["full_fuel_raw"] = str(fuel_val).strip()

        # controls
        rec["controls"]["variable_radiator"] = parse_bool(block.iloc[0, 12]) if block.shape[1] > 12 else None
        rec["controls"]["variable_oil_radiator"] = parse_bool(block.iloc[0, 13]) if block.shape[1] > 13 else None
        rec["controls"]["variable_mixture"] = parse_bool(block.iloc[0, 14]) if block.shape[1] > 14 else None
        rec["controls"]["variable_pitch_prop"] = parse_bool(block.iloc[0, 15]) if block.shape[1] > 15 else None

        rec["engine"]["wep_duration_min"] = clean_num(block.iloc[0, 16]) if block.shape[1] > 16 else None
        rec["engine"]["max_rpm"] = clean_num(block.iloc[0, 17]) if block.shape[1] > 17 else None
        if block.shape[1] > 18:
            tested = block.iloc[0, 18]
            if _is_not_na(tested) and str(tested).strip() not in {"-"}:
                rec["sources"][0]["tested_on_or_before"] = str(tested).strip()

        # altitude curve points
        for r in range(1, len(block)):
            alt = block.iloc[r, 2] if block.shape[1] > 2 else None
            if _is_na(alt):
                continue
            alt_f = _to_float(alt)
            if alt_f is None:
                continue
            t1 = parse_mm_ss_cc(block.iloc[r, 4]) if block.shape[1] > 4 else None
            twep = parse_mm_ss_cc(block.iloc[r, 5]) if block.shape[1] > 5 else None
            speed_kmh = clean_num(block.iloc[r, 6]) if block.shape[1] > 6 else None
            speed_mps = speed_kmh / 3.6 if isinstance(speed_kmh, (int, float)) else None

            rec["performance"]["vy"]["climb_time_curve"].append({"altitude_m": alt_f, "time_setting_1_s": t1, "time_wep_s": twep})
            rec["performance"]["max_speed_wep_curve"].append({"altitude_m": alt_f, "speed_wep_kmh": speed_kmh, "speed_wep_mps": speed_mps})

            perf_points.append({
                "vehicle_id": vehicle_id,
                "vehicle_name": name_str,
                "source_file": os.path.basename(source_file),
                "source_sheet": sheet_name,
                "altitude_m": alt_f,
                "climb_time_setting_1_s": t1,
                "climb_time_wep_s": twep,
                "max_speed_wep_kmh": speed_kmh,
                "max_speed_wep_mps": speed_mps,
            })

        records.append(rec)

    return records, perf_points


def ingest_factchecker2(factchecker2_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(factchecker2_path, read_only=True, data_only=True)
    all_records: List[Dict[str, Any]] = []
    all_points: List[Dict[str, Any]] = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            a1 = ws.cell(1, 1).value
            a2 = ws.cell(2, 1).value
            if str(a1).strip() == 'Aircraft Details' and str(a2).strip() == 'Name':
                df = cast(
                    pd.DataFrame,
                    pd.read_excel(factchecker2_path, sheet_name=sheet, header=None, engine='openpyxl'),
                )
                recs, pts = parse_aircraft_sheet(df, sheet, factchecker2_path)
                all_records.extend(recs)
                all_points.extend(pts)
    finally:
        wb.close()
    return all_records, all_points


# -----------------------
# FactChecker1 (weapons)
# -----------------------

MISSILE_COL_MAP: Dict[str, Tuple[str, FieldCoerce]] = {
    'Initial Mass (kg)': ('mass_initial_kg', float),
    'Mass after boost phase (kg)': ('mass_after_boost_kg', float),
    'Booster Thrust (N)': ('booster_thrust_n', float),
    'Boost time (s)': ('boost_time_s', float),
    'Avg.acc boost phase: (m/s²)': ('boost_avg_accel_mps2', float),
    'Booster ΔV (m/s)': ('boost_delta_v_mps', float),
    'Mass after sustainer phase (kg)': ('mass_after_sustainer_kg', float),
    'Sustainer Thrust (N)': ('sustainer_thrust_n', float),
    'Sustainer time (s)': ('sustainer_time_s', float),
    'Avg.acc sustainer phase: (m/s²)': ('sustainer_avg_accel_mps2', float),
    'Sustainer ΔV (m/s)': ('sustainer_delta_v_mps', float),
    'Total ΔV (m/s)': ('total_delta_v_mps', float),
    'Max Speed (m/s)': ('max_speed_mps', float),
    'Proximity Fuse Range (m)': ('proximity_fuse_range_m', float),
    'Guidance type': ('guidance_type', str),
    'Guidance Delay': ('guidance_delay_s', float),
    'Lock Range (Rear aspect)': ('lock_range_rear_km', float),
    'Lock-on Range (All aspect)  / Pitbull range (FOX3) (km)': ('lock_range_all_aspect_or_pitbull_km', float),
    'Max G-load': ('max_g', float),
    'Time until pull reaches 40%': ('time_to_pull_40pct_s', float),
    'Time until pull reaches 100%': ('time_to_pull_100pct_s', float),
    'Max Launch Angle (FOX1/3) Seeker Limit (FOX2) (deg)': ('max_launch_or_seeker_limit_deg', float),
    'Flare and IRCM detection range (km)': ('flare_ircm_detection_range_km', float),
    'DIRCM detection range: [km]': ('dircm_detection_range_km', float),
    'Head-on lock-on range against afterburning target: [km]': ('headon_lock_range_afterburner_km', float),
    'Un-caged Seeker': ('uncaged_seeker', str),
    'Can be slaved to radar:': ('can_be_slaved_to_radar', str),
    'IRCCM Type': ('irccm_type', str),
    'Seeker FOV (deg)': ('seeker_fov_deg', float),
}

WEAPON_FIELD_PATTERNS: List[Tuple[re.Pattern[str], str, FieldCoerce]] = [
    (re.compile(r'^Mass:\s*\[kg\]\s*$', re.I), 'mass_kg', float),
    (re.compile(r'^Mass at end of booster burn:\s*\[kg\]\s*$', re.I), 'mass_after_boost_kg', float),
    (re.compile(r'^Mass at end of sustainer burn:\s*\[kg\]\s*$', re.I), 'mass_after_sustainer_kg', float),
    (re.compile(r'^Calibre:\s*\[mm\]\s*$', re.I), 'calibre_mm', float),
    (re.compile(r'^Length:\s*\[m\]\s*$', re.I), 'length_m', float),
    (re.compile(r'^Force exerted by booster:\s*\[N\]\s*$', re.I), 'booster_thrust_n', float),
    (re.compile(r'^Burn time of booster:\s*\[s\]\s*$', re.I), 'boost_time_s', float),
    (re.compile(r'^ΔV of booster:\s*\[m/s\]\s*$', re.I), 'boost_delta_v_mps', float),
    (re.compile(r'^Force exerted by sustainer:\s*\[N\]\s*$', re.I), 'sustainer_thrust_n', float),
    (re.compile(r'^Burn time of sustainer:\s*\[s\]\s*$', re.I), 'sustainer_time_s', float),
    (re.compile(r'^ΔV of sustainer:\s*\[m/s\]\s*$', re.I), 'sustainer_delta_v_mps', float),
    (re.compile(r'^Total ΔV:\s*\[m/s\]\s*$', re.I), 'total_delta_v_mps', float),
    (re.compile(r'^Max speed:\s*\[m/s\]\s*$', re.I), 'max_speed_mps', float),
    (re.compile(r'^Explosive mass:\s*\[kg of TNT equivalent\]\s*$', re.I), 'explosive_tnt_eq_kg', float),
    (re.compile(r'^Warhead:\s*$', re.I), 'warhead_type', str),
    (re.compile(r'^Proximity Fuse Range:\s*\[m\]\s*$', re.I), 'proximity_fuse_range_m', float),
    (re.compile(r'^Seeker FOV:\s*\[deg\]\s*$', re.I), 'seeker_fov_deg', float),
    (re.compile(r'^Max G-load:\s*$', re.I), 'max_g', float),
    (re.compile(r'^Guidance type:\s*$', re.I), 'guidance_type', str),
]


def infer_weapon_family_from_sheet(sheet_name: str) -> str:
    s = sheet_name.lower()
    if s.startswith('aam'):
        return 'aam'
    if s.startswith('agm'):
        return 'agm'
    if s.startswith('gbu'):
        return 'gbu'
    if s.startswith('sam'):
        return 'sam'
    if s.startswith('atgm'):
        return 'atgm'
    if s.startswith('ashm'):
        return 'ashm'
    return 'weapon'


def apply_pattern(field_label: str, val: Any) -> Tuple[Optional[str], Any]:
    for pattern, key, typ in WEAPON_FIELD_PATTERNS:
        if pattern.match(field_label):
            v = coerce_value(val)
            if v is None:
                return key, None
            if typ is float:
                try:
                    return key, float(v)
                except Exception:
                    try:
                        return key, float(str(v).replace(',', '').strip())
                    except Exception:
                        return key, None
            return key, str(v).strip()
    return None, None


def parse_transposed_sheet(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    df = cast(
        pd.DataFrame,
        pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl"),
    )
    top = df.iloc[0, 1:]
    non_null_cols = [
        i + 1
        for i, v in enumerate(top)
        if _is_not_na(v) and str(v).strip() not in {"", "nan"}
    ]
    if not non_null_cols:
        return []
    max_col = max(non_null_cols)

    records: List[Dict[str, Any]] = []
    for col in range(1, max_col + 1):
        name = df.iloc[0, col]
        if _is_na(name) or str(name).strip() in {"", "nan", "-"}:
            continue
        name_str = str(name).strip()
        weapon_id = slugify(name_str)
        rec: Dict[str, Any] = {
            "weapon_id": weapon_id,
            "name": name_str,
            "category_sheet": sheet_name,
            "weapon_family": infer_weapon_family_from_sheet(sheet_name),
            "source": {"file": os.path.basename(path), "sheet": sheet_name, "column": col + 1},
            "fields": {},
            "raw_fields": {},
        }

        for row in range(1, len(df)):
            field = df.iloc[row, 0]
            if _is_na(field):
                continue
            field_label = str(field).strip()
            if field_label == '':
                continue
            val = df.iloc[row, col]
            v = coerce_value(val)
            if v is None:
                # likely a section header row; skip
                continue

            rec["raw_fields"][field_label] = v
            k, vv = apply_pattern(field_label, v)
            if k is not None:
                rec["fields"][k] = vv

        records.append(rec)
    return records


def ingest_missile_data(factchecker1_path: str) -> List[Dict[str, Any]]:
    df = cast(
        pd.DataFrame,
        pd.read_excel(factchecker1_path, sheet_name="Missile Data", header=1, engine="openpyxl"),
    )
    df = cast(pd.DataFrame, df.replace({"-": np.nan}))

    records: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        row = cast(pd.Series, row)
        name = row.get('Name')
        if _is_na(name):
            continue
        name_str = str(name).strip()
        weapon_id = slugify(name_str)
        rec: Dict[str, Any] = {
            "weapon_id": weapon_id,
            "name": name_str,
            "weapon_type": "missile",
            "weapon_family": "missile",
            "sources": [{"file": os.path.basename(factchecker1_path), "sheet": "Missile Data"}],
            "fields": {},
            "raw_fields": {},
        }

        # canonical columns
        for col, (key, typ) in MISSILE_COL_MAP.items():
            if col not in row:
                continue
            val = row[col]
            if _is_na(val):
                rec["fields"][key] = None
            else:
                if isinstance(val, np.generic):
                    val = val.item()
                if typ is float:
                    rec["fields"][key] = _to_float(val)
                else:
                    rec["fields"][key] = str(val).strip()

        # preserve other columns
        for col, val in row.items():
            if col == 'Name' or col in MISSILE_COL_MAP:
                continue
            v = coerce_value(val)
            if v is not None:
                rec["raw_fields"][col] = v

        records.append(rec)

    return records


def ingest_factchecker1_weapons_and_sensors(factchecker1_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Returns (weapons, warheads, sensors)."""
    wb = openpyxl.load_workbook(factchecker1_path, read_only=True, data_only=True)

    # Transposed sheets are identified by A1 == 'Name:'
    transposed = []
    try:
        for sheet in wb.sheetnames:
            a1 = wb[sheet].cell(1, 1).value
            if isinstance(a1, str) and a1.strip().startswith('Name:'):
                transposed.append(sheet)
    finally:
        wb.close()

    missile_records = ingest_missile_data(factchecker1_path)

    # build master by weapon_id
    weapons: Dict[str, Dict[str, Any]] = {r['weapon_id']: r for r in missile_records}
    warheads: Dict[str, Dict[str, Any]] = {}

    for sheet in transposed:
        if sheet == 'TGP and sights':
            continue
        for r in parse_transposed_sheet(factchecker1_path, sheet):
            wid = r['weapon_id']
            if wid not in weapons:
                weapons[wid] = {
                    "weapon_id": wid,
                    "name": r['name'],
                    "weapon_type": None,
                    "weapon_family": r.get('weapon_family'),
                    "sources": [],
                    "fields": {},
                    "raw_fields": {},
                }

            weapons[wid].setdefault('sources', []).append(r['source'])
            # merge canonical fields and raw_fields
            weapons[wid]['fields'] = merge_dict(weapons[wid].get('fields', {}), r.get('fields', {}))
            for k, v in r.get('raw_fields', {}).items():
                if k not in weapons[wid].setdefault('raw_fields', {}):
                    weapons[wid]['raw_fields'][k] = v

            # warhead extraction
            if 'explosive_tnt_eq_kg' in r.get('fields', {}) or 'warhead_type' in r.get('fields', {}):
                warhead_id = f"{wid}_warhead"
                wh = warheads.get(warhead_id, {"warhead_id": warhead_id, "weapon_id": wid, "sources": []})
                if 'explosive_tnt_eq_kg' in r['fields']:
                    wh.setdefault('explosive_tnt_eq_kg', r['fields'].get('explosive_tnt_eq_kg'))
                if 'warhead_type' in r['fields']:
                    wh.setdefault('warhead_type', r['fields'].get('warhead_type'))
                if 'proximity_fuse_range_m' in r['fields']:
                    wh.setdefault('proximity_fuse_range_m', r['fields'].get('proximity_fuse_range_m'))
                wh['sources'].append(r['source'])
                warheads[warhead_id] = wh
                weapons[wid]['warhead_id'] = warhead_id

    # Sensors: TGP and sights (also transposed)
    sensors = ingest_factchecker1_sensors(factchecker1_path)

    return list(weapons.values()), list(warheads.values()), sensors


def ingest_factchecker1_sensors(factchecker1_path: str) -> List[Dict[str, Any]]:
    df = cast(
        pd.DataFrame,
        pd.read_excel(factchecker1_path, sheet_name="TGP and sights", header=None, engine="openpyxl"),
    )
    top = df.iloc[0, 1:]
    non_null_cols = [
        i + 1
        for i, v in enumerate(top)
        if _is_not_na(v) and str(v).strip() not in {"", "nan"}
    ]
    if not non_null_cols:
        return []
    max_col = max(non_null_cols)

    records: List[Dict[str, Any]] = []
    for col in range(1, max_col + 1):
        platform = df.iloc[0, col]
        if _is_na(platform) or str(platform).strip() in {"", "nan", "-"}:
            continue
        platform_name = str(platform).strip()
        platform_id = slugify(platform_name)
        rec: Dict[str, Any] = {
            "sensor_package_id": f"{platform_id}_sensors",
            "platform_id": platform_id,
            "platform_name": platform_name,
            "sources": [{"file": os.path.basename(factchecker1_path), "sheet": "TGP and sights", "column": col + 1}],
            "fields": {},
            "raw_fields": {},
        }

        current_section: Optional[str] = None
        for row in range(1, len(df)):
            label = df.iloc[row, 0]
            if _is_na(label):
                continue
            label_str = str(label).strip()
            if label_str == '':
                continue

            val = df.iloc[row, col]
            v = coerce_value(val)

            # section headers are typically like 'Built-in sights' with no per-column values
            if ':' not in label_str and v is None:
                row_vals = [coerce_value(df.iloc[row, c]) for c in range(1, max_col + 1)]
                if all(rv is None for rv in row_vals):
                    current_section = slugify(label_str)
                    continue

            key_base = label_str.rstrip(':').strip()
            key = slugify(key_base)
            full_key = f"{current_section}__{key}" if current_section else key

            rec['raw_fields'][f"{current_section}:{label_str}" if current_section else label_str] = v
            if v is not None:
                rec['fields'][full_key] = v

        records.append(rec)

    return records


def ingest_sea_level_speed(factchecker1_path: str) -> List[Dict[str, Any]]:
    df = cast(
        pd.DataFrame,
        pd.read_excel(
            factchecker1_path,
            sheet_name="Sea level speed & Map Data",
            header=None,
            engine="openpyxl",
        ),
    )
    records: List[Dict[str, Any]] = []
    current_br: Optional[float] = None

    for i in range(2, len(df)):
        br = df.iloc[i, 0]
        name = df.iloc[i, 1]
        mph = df.iloc[i, 2]
        kph = df.iloc[i, 3]
        knots = df.iloc[i, 4]

        if _is_na(name) or _is_na(kph):
            continue
        if _is_not_na(br):
            current_br = _to_float(br)

        name_str = str(name).strip()
        kph_f = _to_float(kph)
        if kph_f is None:
            continue

        rec: Dict[str, Any] = {
            "vehicle_id": slugify(name_str),
            "name": name_str,
            "br": current_br,
            "top_speed_sea_level_kph": kph_f,
            "top_speed_sea_level_mps": kph_f / 3.6,
            "top_speed_sea_level_mph": _to_float(mph) if _is_not_na(mph) else None,
            "top_speed_sea_level_knots": _to_float(knots) if _is_not_na(knots) else None,
            "sources": [{"file": os.path.basename(factchecker1_path), "sheet": "Sea level speed & Map Data", "row": i + 1}],
        }
        records.append(rec)

    return records


# -----------------------
# specialialdata (loadouts)
# -----------------------

def parse_weight_kg(val: Any) -> Optional[float]:
    if val is None or _is_na(val):
        return None
    s = str(val).strip().lower().replace(',', '')
    m = re.match(r'^\(\s*([\d\.]+)\s*kg\s*\)$', s)
    if m:
        return float(m.group(1))
    m = re.match(r'^\s*([\d\.]+)\s*kg\s*$', s)
    if m:
        return float(m.group(1))
    return None


def looks_like_weight_row(values: List[Any]) -> bool:
    nonnull = [v for v in values if _is_not_na(v)]
    if not nonnull:
        return False
    return all(parse_weight_kg(v) is not None for v in nonnull)


def parse_option_string(opt: str) -> Tuple[int, str, Optional[str]]:
    s = str(opt).strip()
    qty = 1
    m = re.match(r'^\s*(\d+)\s*x\s*(.+)$', s, flags=re.I)
    if m:
        qty = int(m.group(1))
        rest = m.group(2).strip()
    else:
        rest = s

    note: Optional[str] = None
    m2 = re.match(r'^(.*)\s*\(([^)]+)\)\s*$', rest)
    if m2:
        base = m2.group(1).strip()
        possible_note = m2.group(2).strip()
        if 'kg' not in possible_note.lower():
            note = possible_note
            rest = base

    return qty, rest, note


def ingest_custom_loadouts(special_path: str, weapon_name_to_id: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    df = cast(
        pd.DataFrame,
        pd.read_excel(special_path, sheet_name="Custom Loadouts", header=None, engine="openpyxl"),
    )

    row0 = cast(pd.Series, df.iloc[0])
    header_positions: List[Tuple[int, str]] = []
    for col in range(df.shape[1]):
        cell = row0[col]
        if _is_not_na(cell) and str(cell).strip() not in {"Custom loadouts"}:
            header_positions.append((col, str(cell).strip()))
    header_positions.sort(key=lambda item: item[0])

    blocks: List[Tuple[int, int, str]] = []
    for i, (start_col, name) in enumerate(header_positions):
        end_col = header_positions[i + 1][0] if i + 1 < len(header_positions) else df.shape[1]
        blocks.append((start_col, end_col, str(name).strip()))

    meta: List[Dict[str, Any]] = []
    options: List[Dict[str, Any]] = []

    for start_col, end_col, aircraft_name in blocks:
        label_col = start_col
        slot_cols = list(range(start_col + 1, end_col))

        # find slot row
        slot_row_idx = None
        for r in range(df.shape[0]):
            if _is_not_na(df.iloc[r, label_col]) and str(df.iloc[r, label_col]).strip().lower() == 'slot':
                slot_row_idx = r
                break
        if slot_row_idx is None:
            continue

        col_to_slot: Dict[int, int] = {}
        for c in slot_cols:
            v = df.iloc[slot_row_idx, c]
            if _is_na(v):
                continue
            slot_val = _to_int(v)
            if slot_val is not None:
                col_to_slot[c] = slot_val

        mrec: Dict[str, Any] = {
            "vehicle_name": aircraft_name,
            "vehicle_id": slugify(aircraft_name),
            "sources": [{"file": os.path.basename(special_path), "sheet": "Custom Loadouts", "start_col": start_col + 1}],
        }

        max_load_str = df.iloc[1, label_col]
        if _is_not_na(max_load_str):
            m = re.search(r'max load\s*[--]\s*([\d\.]+)\s*kg', str(max_load_str).lower().replace(',', ''))
            if m:
                mrec['max_load_kg'] = float(m.group(1))

        max_left_str = df.iloc[2, label_col]
        if _is_not_na(max_left_str):
            m = re.search(r'max left load\s*[--]\s*([\d\.]+)\s*kg', str(max_left_str).lower().replace(',', ''))
            if m:
                mrec['max_left_load_kg'] = float(m.group(1))

        exempt_label_row = None
        for r in range(df.shape[0]):
            if _is_not_na(df.iloc[r, label_col]) and str(df.iloc[r, label_col]).strip().lower().startswith('exempt'):
                exempt_label_row = r
                break
        if exempt_label_row is not None:
            exempt_val = df.iloc[exempt_label_row, start_col + 1] if start_col + 1 < end_col else None
            mrec['exempt_from_imbalance_calcs'] = parse_bool(exempt_val)

        meta.append(mrec)

        # category rows
        category_rows: List[Tuple[int, str]] = []
        for r in range(slot_row_idx + 1, df.shape[0]):
            label = df.iloc[r, label_col]
            if _is_na(label):
                continue
            label_str = str(label).strip()
            if label_str == '':
                continue
            if label_str.lower().startswith('max ') or label_str.lower().startswith('exempt') or label_str.lower() == 'slot':
                continue
            category_rows.append((r, label_str))
        category_rows = sorted(category_rows, key=lambda x: x[0])

        for idx, (cat_row, cat_name) in enumerate(category_rows):
            next_row = category_rows[idx + 1][0] if idx + 1 < len(category_rows) else df.shape[0]
            pending: Dict[int, List[int]] = {}

            for r in range(cat_row, next_row):
                row_vals = [df.iloc[r, c] for c in col_to_slot.keys()]
                if all(_is_na(v) for v in row_vals):
                    continue

                if looks_like_weight_row(row_vals):
                    for c, slot in col_to_slot.items():
                        wval = df.iloc[r, c]
                        if _is_na(wval):
                            continue
                        wkg = parse_weight_kg(wval)
                        if c in pending:
                            for idx_opt in pending[c]:
                                options[idx_opt]['mass_kg'] = wkg
                    pending = {}
                else:
                    pending = {}
                    for c, slot in col_to_slot.items():
                        oval = df.iloc[r, c]
                        if _is_na(oval):
                            continue
                        opt_str = str(oval).strip()
                        if opt_str in {'', '-'}:
                            continue

                        qty, weapon_name, note = parse_option_string(opt_str)
                        weapon_id_match = weapon_name_to_id.get(slugify(weapon_name))

                        orec: Dict[str, Any] = {
                            "vehicle_id": mrec['vehicle_id'],
                            "vehicle_name": aircraft_name,
                            "slot": slot,
                            "category": cat_name,
                            "option": opt_str,
                            "quantity": qty,
                            "weapon_name": weapon_name,
                            "option_note": note,
                            "weapon_id_match": weapon_id_match,
                            "mass_kg": None,
                            "sources": [{"file": os.path.basename(special_path), "sheet": "Custom Loadouts", "row": r + 1, "col": c + 1}],
                        }
                        options.append(orec)
                        pending.setdefault(c, []).append(len(options) - 1)

    # alias candidates: early/late -> base
    aliases: List[Dict[str, Any]] = []
    for mrec in meta:
        name = mrec['vehicle_name']
        base = re.sub(r'\((early|late)\)', '', name, flags=re.I).strip()
        base = re.sub(r'\b(early|late)\b$', '', base, flags=re.I).strip()
        base_vid = slugify(base)
        if base_vid != mrec['vehicle_id']:
            aliases.append({
                "alias_vehicle_id": mrec['vehicle_id'],
                "alias_name": name,
                "canonical_vehicle_id": base_vid,
                "canonical_name": base,
                "rule": "strip early/late",
            })

    return meta, options, aliases


def write_jsonl(path: Path, records: List[Dict[str, Any]]) -> None:
    with path.open('w', encoding='utf-8') as f:
        for rec in records:
            f.write(json.dumps(make_json_safe(rec), ensure_ascii=False) + '\n')


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--factchecker1', required=True)
    ap.add_argument('--factchecker2', required=True)
    ap.add_argument('--factchecker3', required=False)  # reserved for later
    ap.add_argument('--special', required=True)
    ap.add_argument('--out_dir', required=True)
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # FactChecker2 aircraft
    fc2_records, fc2_points = ingest_factchecker2(args.factchecker2)

    # FactChecker1 weapons + sensors
    weapons, warheads, sensors = ingest_factchecker1_weapons_and_sensors(args.factchecker1)

    # Build a lookup for loadout matching (name->weapon_id)
    weapon_name_to_id = {slugify(w['name']): w['weapon_id'] for w in weapons if w.get('name')}

    # specialialdata loadouts
    loadout_meta, loadout_options, alias_candidates = ingest_custom_loadouts(args.special, weapon_name_to_id)

    # sea level speed
    sea_level = ingest_sea_level_speed(args.factchecker1)

    # vehicles: merge across sources
    vehicles: Dict[str, Dict[str, Any]] = {v['vehicle_id']: v for v in fc2_records}

    # merge sea-level
    for v in sea_level:
        vid = v['vehicle_id']
        if vid not in vehicles:
            vehicles[vid] = {
                "vehicle_id": vid,
                "name": v['name'],
                "vehicle_type": "aircraft",
                "notes": None,
                "engine": {},
                "supercharger": {},
                "controls": {},
                "fuel": {},
                "performance": {},
                "sources": [],
            }
        vehicles[vid].setdefault('performance', {}).setdefault('sea_level', {})
        vehicles[vid]['performance']['sea_level'].update({
            'top_speed_kph': v['top_speed_sea_level_kph'],
            'top_speed_mps': v['top_speed_sea_level_mps'],
            'br': v.get('br'),
        })
        vehicles[vid].setdefault('sources', []).extend(v.get('sources', []))

    # merge sensors
    for s in sensors:
        vid = s['platform_id']
        if vid not in vehicles:
            vehicles[vid] = {
                "vehicle_id": vid,
                "name": s['platform_name'],
                "vehicle_type": "aircraft",
                "notes": None,
                "engine": {},
                "supercharger": {},
                "controls": {},
                "fuel": {},
                "performance": {},
                "sources": [],
            }
        vehicles[vid].setdefault('sensor_package_ids', []).append(s['sensor_package_id'])
        vehicles[vid].setdefault('sources', []).extend(s.get('sources', []))

    # merge loadout meta
    for m in loadout_meta:
        vid = m['vehicle_id']
        if vid not in vehicles:
            vehicles[vid] = {
                "vehicle_id": vid,
                "name": m['vehicle_name'],
                "vehicle_type": "aircraft",
                "notes": None,
                "engine": {},
                "supercharger": {},
                "controls": {},
                "fuel": {},
                "performance": {},
                "sources": [],
            }
        vehicles[vid]['custom_loadout_meta'] = {k: v for k, v in m.items() if k not in {'vehicle_id', 'vehicle_name', 'sources'}}
        vehicles[vid].setdefault('sources', []).extend(m.get('sources', []))

    # Write outputs
    write_jsonl(out_dir / 'vehicles.jsonl', sorted(vehicles.values(), key=lambda x: x.get('vehicle_id', '')))

    pd.DataFrame(fc2_points).to_csv(out_dir / 'aircraft_performance_points_fc2.csv', index=False)
    pd.DataFrame(sea_level).to_csv(out_dir / 'sea_level_speed_fc1.csv', index=False)

    write_jsonl(out_dir / 'weapons.jsonl', sorted(weapons, key=lambda x: x.get('weapon_id', '')))
    write_jsonl(out_dir / 'warheads.jsonl', sorted(warheads, key=lambda x: x.get('warhead_id', '')))
    write_jsonl(out_dir / 'sensors_tgp_and_sights.jsonl', sorted(sensors, key=lambda x: x.get('sensor_package_id', '')))

    write_jsonl(out_dir / 'custom_loadout_meta.jsonl', sorted(loadout_meta, key=lambda x: x.get('vehicle_id', '')))
    write_jsonl(out_dir / 'custom_loadout_options.jsonl', sorted(loadout_options, key=lambda x: (x.get('vehicle_id', ''), x.get('slot', -1), x.get('category', ''), x.get('option', ''))))

    write_jsonl(out_dir / 'vehicle_alias_candidates.jsonl', alias_candidates)

    print('Done.')
    print(f'Wrote outputs to: {out_dir.resolve()}')


if __name__ == '__main__':
    main()
