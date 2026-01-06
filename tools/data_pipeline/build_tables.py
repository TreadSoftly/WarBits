from __future__ import annotations

import json
import re
import unicodedata
from datetime import time as time_type
from pathlib import Path
from typing import Any, Callable, Iterator, cast, TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Worksheet = Any

RowValues = tuple[Any, ...]
RowMap = dict[str, RowValues]

ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT / "warbits" / "data"
RAW_DIR = ROOT / "warbits" / "data" / "raw"

def _raw_file(name: str) -> Path:
    candidate = RAW_DIR / name
    if candidate.exists():
        return candidate
    return ROOT / name


FACTCHECKER1 = _raw_file("WarThunder_FactChecker1.xlsx")
FACTCHECKER2 = _raw_file("WarThunder_FactChecker2.xlsx")
SPECIALIAL = _raw_file("specialialdata.xlsx")

WEAPON_PREFIXES = ("AAM", "AGM", "GBU", "SAM", "ATGM", "AShM")
SKIP_SHEET_PREFIXES = ("Copy of",)
WEAPON_SECTION_HEADERS = {
    "physical_properties",
    "engine_properties",
    "fuse_and_warhead_properties",
    "guidance_properties",
    "flight_characteristics",
}


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower()
    ascii_text = re.sub(r"[^a-z0-9]+", "_", ascii_text)
    return ascii_text.strip("_")


def normalize_label(label: str) -> str:
    label = label.replace(chr(0x0394), "Delta ")
    label = re.sub(r"\[[^\]]+\]", "", label)
    label = label.replace(":", " ")
    label = unicodedata.normalize("NFKD", label)
    label = label.encode("ascii", "ignore").decode("ascii")
    label = label.lower()
    label = re.sub(r"[^a-z0-9]+", "_", label)
    return label.strip("_")


def try_parse_float(value: str) -> float | None:
    raw = value.strip()
    if not raw:
        return None
    if "/" in raw:
        return None
    raw = raw.replace(",", ".")
    raw = raw.strip()
    raw = raw.rstrip("x%")
    if not raw:
        return None
    if re.search(r"[a-zA-Z]", raw):
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return None
        lowered = text.lower()
        if lowered in {"yes", "no"}:
            return lowered == "yes"
        if len(text) == 1:
            code = ord(text)
            if code == 0x2714:
                return True
            if code == 0x2718:
                return False
        numeric = try_parse_float(text)
        if numeric is not None:
            return numeric
        return text
    if isinstance(value, (int, float, bool)):
        return value
    return value


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return try_parse_float(value)
    return None


def to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"yes", "true", "1"}:
            return True
        if lowered in {"no", "false", "0"}:
            return False
        if len(value) == 1:
            code = ord(value)
            if code == 0x2714:
                return True
            if code == 0x2718:
                return False
    return None


def time_to_seconds(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, time_type):
        if value.second == 0:
            return float(value.hour * 60 + value.minute)
        return float(value.hour * 3600 + value.minute * 60 + value.second)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        parts = text.split(":")
        if len(parts) == 2 and all(p.isdigit() for p in parts):
            return float(int(parts[0]) * 60 + int(parts[1]))
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            return float(int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2]))
    return None


def parse_duration_minutes(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().lower()
        if not text:
            return None
        hours = 0
        minutes = 0
        hours_match = re.search(r"(\d+)\s*h", text)
        minutes_match = re.search(r"(\d+)\s*m", text)
        if hours_match:
            hours = int(hours_match.group(1))
        if minutes_match:
            minutes = int(minutes_match.group(1))
        if hours_match or minutes_match:
            return float(hours * 60 + minutes)
    return None


def parse_range(value: Any) -> dict[str, float] | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        return {"min": num, "max": num}
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return None
        text = text.replace(chr(0x00D7), "x")
        text = text.replace(",", ".")
        if "/" in text:
            parts = [p.strip() for p in text.split("/")]
            if len(parts) == 2:
                left = try_parse_float(parts[0])
                right = try_parse_float(parts[1])
                if left is not None and right is not None:
                    return {"min": left, "max": right}
    return None


def parse_resolution(value: Any) -> dict[str, int] | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip().lower()
        if not text or text == "-":
            return None
        text = text.replace(chr(0x00D7), "x")
        text = text.replace(" ", "")
        parts = [p for p in text.split("x") if p]
        if len(parts) == 2 and all(p.isdigit() for p in parts):
            return {"width": int(parts[0]), "height": int(parts[1])}
    return None


def iter_values(ws: Worksheet, **kwargs: Any) -> Iterator[RowValues]:
    return cast(Iterator[RowValues], ws.iter_rows(values_only=True, **kwargs))


def get_row_value(row_map: RowMap, key: str, idx: int) -> Any:
    row = row_map.get(key)
    if not row or idx >= len(row):
        return None
    return row[idx]


def add_attribute(target: dict[str, Any], key: str, value: Any) -> None:
    if value is None:
        return
    if key in target:
        existing = target[key]
        if isinstance(existing, list):
            cast(list[Any], existing).append(value)  # type: ignore[redundant-cast]
        else:
            target[key] = [existing, value]
        return
    target[key] = value


def weapon_type_from_sheet(sheet_name: str) -> tuple[str, str | None]:
    name = sheet_name.strip()
    if name.startswith("AAM"):
        return "air_to_air_missile", _sheet_tags(sheet_name)
    if name.startswith("AGM"):
        return "air_to_ground_missile", _sheet_tags(sheet_name)
    if name.startswith("GBU"):
        return "guided_bomb", _sheet_tags(sheet_name)
    if name.startswith("SAM"):
        return "surface_to_air_missile", _sheet_tags(sheet_name)
    if name.startswith("ATGM"):
        return "anti_tank_guided_missile", _sheet_tags(sheet_name)
    if name.startswith("AShM"):
        return "anti_ship_missile", _sheet_tags(sheet_name)
    return "weapon", _sheet_tags(sheet_name)


def _sheet_tags(sheet_name: str) -> str | None:
    tags = re.findall(r"\(([^)]+)\)", sheet_name)
    if not tags:
        return None
    return ", ".join(tag.strip() for tag in tags if tag.strip())


def scale_float(value: Any, scale: float) -> float | None:
    raw = to_float(value)
    if raw is None:
        return None
    return raw * scale


def mm_to_m(value: Any) -> float | None:
    return scale_float(value, 0.001)


def km_to_m(value: Any) -> float | None:
    return scale_float(value, 1000.0)


WEAPON_FIELD_MAP: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "mass": ("mass_kg", to_float),
    "mass_at_end_of_booster_burn": ("mass_booster_end_kg", to_float),
    "mass_at_end_of_sustainer_burn": ("mass_sustainer_end_kg", to_float),
    "calibre": ("diameter_m", mm_to_m),
    "length": ("length_m", to_float),
    "force_exerted_by_booster": ("booster_thrust_n", to_float),
    "burn_time_of_booster": ("booster_time_s", to_float),
    "force_exerted_by_sustainer": ("sustainer_thrust_n", to_float),
    "burn_time_of_sustainer": ("sustainer_time_s", to_float),
    "maximum_speed": ("max_speed_mps", to_float),
    "minimum_range": ("min_range_m", to_float),
    "flight_range_limit": ("max_range_m", km_to_m),
    "guidance_type": ("guidance_type", lambda v: str(v).strip() if v is not None else None),
    "field_of_view": ("fov_deg", to_float),
    "gimbal_limit": ("gimbal_limit_deg", to_float),
    "track_rate": ("track_rate_deg_s", to_float),
    "guidance_start_delay": ("guidance_start_delay_s", to_float),
    "guidance_duration": ("guidance_duration_s", to_float),
    "seeker_warm_up_time": ("seeker_warmup_s", to_float),
}

WARHEAD_FIELD_MAP: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "explosive_mass": ("explosive_mass_kg", to_float),
    "warhead": ("warhead_type", lambda v: str(v).strip() if v is not None else None),
    "penetration": ("penetration_mm", to_float),
    "proximity_fuse": ("proximity_fuse", to_bool),
    "proximity_fuse_range": ("proximity_fuse_range_m", to_float),
    "proximity_fuse_delay": ("proximity_fuse_delay_s", to_float),
    "impact_fuse_sensitivity": ("impact_fuse_sensitivity_mm", to_float),
    "impact_fuse_delay": ("impact_fuse_delay_m", to_float),
}


def extract_weapons(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    weapons_by_id: dict[str, dict[str, Any]] = {}
    warheads_by_id: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith(SKIP_SHEET_PREFIXES):
                continue
            if not sheet_name.startswith(WEAPON_PREFIXES):
                continue
            ws = wb[sheet_name]
            header = next(iter_values(ws, min_row=1, max_row=1), None)
            if not header or not header[0] or "Name" not in str(header[0]):
                warnings.append(f"Skipping {sheet_name}: missing Name header.")
                continue

            column_names: dict[int, str] = {}
            for col_idx, name in enumerate(header, start=1):
                if col_idx == 1:
                    continue
                cleaned = clean_value(name)
                if not cleaned:
                    continue
                column_names[col_idx] = str(cleaned)

            for col_idx, weapon_name in column_names.items():
                weapon_id = slugify(weapon_name)
                if weapon_id in weapons_by_id:
                    sources = cast(
                        list[dict[str, str]],
                        weapons_by_id[weapon_id].get("sources", []),
                    )
                    sources.append({"book": path.name, "sheet": sheet_name})
                    weapons_by_id[weapon_id]["sources"] = sources
                    continue
                weapon_type, sheet_tag = weapon_type_from_sheet(sheet_name)
                weapon: dict[str, Any] = {
                    "id": weapon_id,
                    "name": weapon_name,
                    "weapon_type": weapon_type,
                    "weapon_tag": sheet_tag,
                    "sources": [{"book": path.name, "sheet": sheet_name}],
                    "attributes": {},
                }
                warhead_id = f"warhead_{weapon_id}"
                warhead: dict[str, Any] = {
                    "id": warhead_id,
                    "weapon_id": weapon_id,
                    "sources": [{"book": path.name, "sheet": sheet_name}],
                    "attributes": {},
                }
                weapon["warhead_id"] = warhead_id
                weapons_by_id[weapon_id] = weapon
                warheads_by_id[warhead_id] = warhead

            for row in iter_values(ws, min_row=2):
                label = row[0]
                if not label:
                    continue
                label_key = normalize_label(str(label))
                if label_key in WEAPON_SECTION_HEADERS:
                    continue

                for col_idx, weapon_name in column_names.items():
                    value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                    cleaned = clean_value(value)
                    if cleaned is None:
                        continue
                    weapon_id = slugify(weapon_name)
                    weapon = weapons_by_id[weapon_id]
                    warhead = warheads_by_id[f"warhead_{weapon_id}"]

                    if label_key in WEAPON_FIELD_MAP:
                        field, converter = WEAPON_FIELD_MAP[label_key]
                        parsed = converter(cleaned)
                        if parsed is not None:
                            weapon[field] = parsed
                        continue
                    if label_key in WARHEAD_FIELD_MAP:
                        field, converter = WARHEAD_FIELD_MAP[label_key]
                        parsed = converter(cleaned)
                        if parsed is not None:
                            warhead[field] = parsed
                        continue

                    add_attribute(weapon["attributes"], label_key, cleaned)

    finally:
        wb.close()

    return list(weapons_by_id.values()), list(warheads_by_id.values()), warnings


def is_vehicle_sheet(ws: Worksheet) -> bool:
    header = ws.cell(row=1, column=1).value
    name_header = ws.cell(row=2, column=1).value
    return str(header).strip().lower() == "aircraft details" and str(name_header).strip().lower() == "name"


def parse_vehicle_sheet(
    ws: Worksheet,
    sheet_name: str,
    vehicles_by_id: dict[str, dict[str, Any]],
) -> None:
    current: dict[str, Any] | None = None
    profile: list[dict[str, Any]] = []

    for row in iter_values(ws, min_row=3):
        name = row[0]
        if isinstance(name, str) and name.strip():
            if name.strip().lower() in {"name", "aircraft details"}:
                continue
            if current:
                finalize_vehicle(current, profile, vehicles_by_id)
                profile = []
            current = build_vehicle_base(row, sheet_name)
            continue

        if not current:
            continue

        altitude = row[2] if len(row) > 2 else None
        if isinstance(altitude, (int, float)):
            entry: dict[str, Any] = {
                "altitude_m": float(altitude),
                "time_nominal_s": time_to_seconds(row[4] if len(row) > 4 else None),
                "time_wep_s": time_to_seconds(row[5] if len(row) > 5 else None),
                "max_speed_wep_kmh": to_float(row[6] if len(row) > 6 else None),
            }
            profile.append(entry)

    if current:
        finalize_vehicle(current, profile, vehicles_by_id)


def build_vehicle_base(row: tuple[Any, ...], sheet_name: str) -> dict[str, Any]:
    name = str(row[0]).strip()
    notes = clean_value(row[1])
    stage_2 = to_float(row[7])
    stage_3 = to_float(row[8])
    stage_2_hp = to_float(row[9])
    stage_3_hp = to_float(row[10])
    fuel_minutes = parse_duration_minutes(row[11])
    variable_rad = to_bool(row[12])
    variable_oil = to_bool(row[13])

    return {
        "id": slugify(name),
        "name": name,
        "vehicle_type": "aircraft",
        "notes": notes,
        "supercharger_stage_2_m": stage_2,
        "supercharger_stage_3_m": stage_3,
        "engine_power_stage_2_hp": stage_2_hp,
        "engine_power_stage_3_hp": stage_3_hp,
        "fuel_full_duration_min": fuel_minutes,
        "variable_radiator": variable_rad,
        "variable_oil_radiator": variable_oil,
        "sources": [{"book": FACTCHECKER2.name, "sheet": sheet_name}],
    }


def compute_best_climb(profile: list[dict[str, Any]], key: str) -> tuple[float | None, float | None]:
    sorted_profile = sorted(
        [p for p in profile if p.get(key) is not None],
        key=lambda item: item["altitude_m"],
    )
    best_rate = None
    best_alt = None
    last_alt = None
    last_time = None
    for point in sorted_profile:
        alt = point["altitude_m"]
        time_value = point[key]
        if last_alt is None or last_time is None:
            last_alt = alt
            last_time = time_value
            continue
        delta_alt = alt - last_alt
        delta_time = time_value - last_time
        if delta_time > 0 and delta_alt > 0:
            rate = delta_alt / delta_time
            if best_rate is None or rate > best_rate:
                best_rate = rate
                best_alt = alt
        last_alt = alt
        last_time = time_value
    return best_rate, best_alt


def finalize_vehicle(
    vehicle: dict[str, Any],
    profile: list[dict[str, Any]],
    vehicles_by_id: dict[str, dict[str, Any]],
) -> None:
    if profile:
        vehicle["climb_profile"] = profile
        max_speed = None
        max_speed_alt = None
        for point in profile:
            speed = point.get("max_speed_wep_kmh")
            if speed is None:
                continue
            if max_speed is None or speed > max_speed:
                max_speed = speed
                max_speed_alt = point.get("altitude_m")
        if max_speed is not None:
            vehicle["max_speed_mps"] = max_speed / 3.6
            vehicle["max_speed_altitude_m"] = max_speed_alt

        best_rate, best_alt = compute_best_climb(profile, "time_wep_s")
        if best_rate is None:
            best_rate, best_alt = compute_best_climb(profile, "time_nominal_s")
        if best_rate is not None:
            vehicle["best_climb_rate_mps"] = best_rate
            vehicle["best_climb_altitude_m"] = best_alt

    vehicle_id = vehicle["id"]
    if vehicle_id not in vehicles_by_id:
        vehicles_by_id[vehicle_id] = vehicle
        return

    existing = vehicles_by_id[vehicle_id]
    sources = cast(list[dict[str, str]], existing.get("sources", []))
    extra_sources = cast(list[dict[str, str]], vehicle.get("sources", []))
    sources.extend(extra_sources)
    existing["sources"] = sources
    for key, value in vehicle.items():
        if key in {"sources"}:
            continue
        if existing.get(key) is None and value is not None:
            existing[key] = value


def extract_vehicles(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    vehicles_by_id: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not is_vehicle_sheet(ws):
                continue
            parse_vehicle_sheet(ws, sheet_name, vehicles_by_id)
    finally:
        wb.close()
    if not vehicles_by_id:
        warnings.append("No vehicle sheets parsed.")
    return list(vehicles_by_id.values()), warnings


def extract_sensors(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    sensors: list[dict[str, Any]] = []
    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "TGP and sights" not in wb.sheetnames:
            warnings.append("Missing TGP and sights sheet.")
            return sensors, warnings
        ws = wb["TGP and sights"]
        rows = list(iter_values(ws))
        if not rows:
            warnings.append("TGP and sights sheet is empty.")
            return sensors, warnings

        platform_row = rows[0]
        platform_names = [clean_value(name) for name in platform_row[1:]]

        built_in_start = None
        pod_start = None
        for idx, row in enumerate(rows):
            label = row[0]
            if not label:
                continue
            key = normalize_label(str(label))
            if key == "built_in_sights":
                built_in_start = idx
            if key == "targeting_pod":
                pod_start = idx
        if built_in_start is None or pod_start is None:
            warnings.append("TGP and sights sheet missing expected sections.")
            return sensors, warnings

        built_in_rows = rows[built_in_start + 1 : pod_start]
        pod_rows = rows[pod_start + 1 :]

        built_in_map: RowMap = {}
        for row in built_in_rows:
            if not row or not row[0]:
                continue
            built_in_map[normalize_label(str(row[0]))] = row
        pod_map: RowMap = {}
        for row in pod_rows:
            if not row or not row[0]:
                continue
            pod_map[normalize_label(str(row[0]))] = row

        for idx, platform in enumerate(platform_names, start=1):
            if not platform:
                continue
            platform_name = str(platform)

            built_in: dict[str, Any] = {
                "id": slugify(f"{platform_name}_built_in_sight"),
                "platform_name": platform_name,
                "role": "built_in_sight",
                "sensor_type": clean_value(get_row_value(built_in_map, "type", idx)),
                "horizontal_limits_deg": parse_range(
                    get_row_value(built_in_map, "horizontal_limits", idx)
                ),
                "vertical_limits_deg": parse_range(
                    get_row_value(built_in_map, "vertical_limits", idx)
                ),
                "zoom_range": parse_range(get_row_value(built_in_map, "zoom", idx)),
                "thermal_resolution": parse_resolution(
                    get_row_value(built_in_map, "thermal_resolution", idx)
                ),
                "nvd_resolution": parse_resolution(
                    get_row_value(built_in_map, "nvd_resolution", idx)
                ),
                "sources": [{"book": path.name, "sheet": "TGP and sights"}],
            }
            if any(
                built_in[field] is not None
                for field in (
                    "sensor_type",
                    "horizontal_limits_deg",
                    "vertical_limits_deg",
                    "zoom_range",
                    "thermal_resolution",
                    "nvd_resolution",
                )
            ):
                sensors.append(built_in)

            pod_name = clean_value(get_row_value(pod_map, "name", idx))
            pod_entry: dict[str, Any] = {
                "id": slugify(f"{platform_name}_targeting_pod"),
                "platform_name": platform_name,
                "role": "targeting_pod",
                "name": pod_name,
                "horizontal_limits_deg": parse_range(
                    get_row_value(pod_map, "horizontal_limits", idx)
                ),
                "vertical_limits_deg": parse_range(
                    get_row_value(pod_map, "vertical_limits", idx)
                ),
                "zoom_range": parse_range(get_row_value(pod_map, "zoom", idx)),
                "thermal_resolution": parse_resolution(
                    get_row_value(pod_map, "thermal_resolution", idx)
                ),
                "nvd_resolution": parse_resolution(
                    get_row_value(pod_map, "nvd_resolution", idx)
                ),
                "sources": [{"book": path.name, "sheet": "TGP and sights"}],
            }
            if any(
                pod_entry[field] is not None
                for field in (
                    "name",
                    "horizontal_limits_deg",
                    "vertical_limits_deg",
                    "zoom_range",
                    "thermal_resolution",
                    "nvd_resolution",
                )
            ):
                sensors.append(pod_entry)
    finally:
        wb.close()

    return sensors, warnings


def default_terrain() -> list[dict[str, Any]]:
    return [
        {
            "id": "desert_sand",
            "name": "Desert sand",
            "biome": "desert",
            "surface": "sand",
            "friction_coefficient": 0.4,
            "roughness": 0.6,
            "notes": "Seed values; replace with measured data.",
        },
        {
            "id": "urban_asphalt",
            "name": "Urban asphalt",
            "biome": "urban",
            "surface": "asphalt",
            "friction_coefficient": 0.9,
            "roughness": 0.2,
            "notes": "Seed values; replace with measured data.",
        },
        {
            "id": "mountain_rock",
            "name": "Mountain rock",
            "biome": "mountain",
            "surface": "rock",
            "friction_coefficient": 0.7,
            "roughness": 0.8,
            "notes": "Seed values; replace with measured data.",
        },
        {
            "id": "forest_loam",
            "name": "Forest loam",
            "biome": "forest",
            "surface": "soil",
            "friction_coefficient": 0.6,
            "roughness": 0.7,
            "notes": "Seed values; replace with measured data.",
        },
    ]


def default_loadouts() -> dict[str, Any]:
    return {
        "loadouts": [],
        "hardpoints": [],
        "loadout_items": [],
        "notes": (
            "Custom loadouts sheet is irregular; defer parsing until schema is fixed."
        ),
        "sources": [{"book": SPECIALIAL.name, "sheet": "Custom Loadouts"}],
    }


def validate_tables(
    vehicles: list[dict[str, Any]],
    weapons: list[dict[str, Any]],
    warheads: list[dict[str, Any]],
    sensors: list[dict[str, Any]],
    terrain: list[dict[str, Any]],
    loadouts: dict[str, Any],
) -> list[str]:
    warnings: list[str] = []

    def check_unique(entries: list[dict[str, Any]], label: str) -> None:
        seen: set[str] = set()
        for entry in entries:
            entry_id = entry.get("id")
            if not entry_id:
                warnings.append(f"{label} entry missing id.")
                continue
            if entry_id in seen:
                warnings.append(f"Duplicate {label} id: {entry_id}")
            seen.add(entry_id)

    check_unique(vehicles, "vehicle")
    check_unique(weapons, "weapon")
    check_unique(warheads, "warhead")
    check_unique(sensors, "sensor")
    check_unique(terrain, "terrain")

    warhead_ids = {w["id"] for w in warheads if w.get("id")}
    for weapon in weapons:
        warhead_id = weapon.get("warhead_id")
        if warhead_id and warhead_id not in warhead_ids:
            warnings.append(
                f"Weapon {weapon.get('id')} references missing warhead {warhead_id}"
            )

    if not loadouts.get("loadouts"):
        warnings.append("Loadouts table is empty.")

    return warnings


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(
            payload, indent=2, ensure_ascii=True, sort_keys=True, default=str
        )
    )


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    weapons, warheads, weapon_warnings = extract_weapons(FACTCHECKER1)
    vehicles, vehicle_warnings = extract_vehicles(FACTCHECKER2)
    sensors, sensor_warnings = extract_sensors(FACTCHECKER1)
    terrain = default_terrain()
    loadouts = default_loadouts()

    validation_warnings = validate_tables(
        vehicles, weapons, warheads, sensors, terrain, loadouts
    )

    write_json(OUTPUT_DIR / "vehicles.json", vehicles)
    write_json(OUTPUT_DIR / "weapons.json", weapons)
    write_json(OUTPUT_DIR / "warheads.json", warheads)
    write_json(OUTPUT_DIR / "sensors.json", sensors)
    write_json(OUTPUT_DIR / "terrain.json", terrain)
    write_json(OUTPUT_DIR / "loadouts.json", loadouts)

    summary: dict[str, Any] = {
        "counts": {
            "vehicles": len(vehicles),
            "weapons": len(weapons),
            "warheads": len(warheads),
            "sensors": len(sensors),
            "terrain": len(terrain),
        },
        "sources": {
            "vehicles": FACTCHECKER2.name,
            "weapons": FACTCHECKER1.name,
            "warheads": FACTCHECKER1.name,
            "sensors": FACTCHECKER1.name,
            "loadouts": SPECIALIAL.name,
        },
        "warnings": weapon_warnings + vehicle_warnings + sensor_warnings + validation_warnings,
    }
    write_json(OUTPUT_DIR / "data_summary.json", summary)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
